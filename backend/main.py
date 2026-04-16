from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import or_
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import json
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Processo, get_db, create_tables
from datajud import buscar_todos_processos

app = FastAPI(title="Sistema Jurídico - Daniele Cabral", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

sincronizacao_status = {"em_andamento": False, "ultima_vez": None, "total": 0, "erros": []}


@app.on_event("startup")
async def startup():
    create_tables()


# Servir frontend (após todos os endpoints /api serem registrados)
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")


def montar_frontend():
    if os.path.isdir(STATIC_DIR):
        app.mount("/", StaticFiles(directory=STATIC_DIR, html=True), name="frontend")


# ─── Modelos Pydantic ───────────────────────────────────────────────────────

class AtualizacaoProcesso(BaseModel):
    o_que_fazer: Optional[str] = None
    prioridade: Optional[str] = None
    oculto: Optional[bool] = None


class RespostaProcesso(BaseModel):
    numero: str
    tribunal: Optional[str]
    vara: Optional[str]
    classe: Optional[str]
    assunto: Optional[str]
    data_distribuicao: Optional[str]
    valor_causa: Optional[float]
    situacao: Optional[str]
    ultimo_movimento: Optional[str]
    data_ultimo_movimento: Optional[str]
    polo_ativo: Optional[str]
    polo_passivo: Optional[str]
    o_que_fazer: Optional[str]
    prioridade: str
    atualizado_em: Optional[datetime]
    criado_em: Optional[datetime]
    oculto: bool
    fonte_oab: Optional[str]

    class Config:
        from_attributes = True


# ─── Sincronização ──────────────────────────────────────────────────────────

async def _executar_sincronizacao(db: Session):
    sincronizacao_status["em_andamento"] = True
    sincronizacao_status["erros"] = []
    try:
        processos, erros = await buscar_todos_processos()
        sincronizacao_status["erros"] = erros

        total_novos = 0
        for dados in processos:
            existente = db.query(Processo).filter(Processo.numero == dados["numero"]).first()
            if existente:
                # Atualiza apenas campos do tribunal, preserva anotações do usuário
                for campo in ["tribunal", "vara", "classe", "assunto", "data_distribuicao",
                              "valor_causa", "situacao", "ultimo_movimento",
                              "data_ultimo_movimento", "polo_ativo", "polo_passivo", "partes", "fonte_oab"]:
                    if dados.get(campo) is not None:
                        setattr(existente, campo, dados[campo])
                existente.atualizado_em = datetime.utcnow()
            else:
                novo = Processo(**dados)
                db.add(novo)
                total_novos += 1

        db.commit()
        sincronizacao_status["total"] = db.query(Processo).count()
        sincronizacao_status["ultima_vez"] = datetime.now().isoformat()
    except Exception as e:
        sincronizacao_status["erros"].append(str(e))
    finally:
        sincronizacao_status["em_andamento"] = False


# ─── Endpoints ──────────────────────────────────────────────────────────────

@app.get("/api/status")
def status_sistema(db: Session = Depends(get_db)):
    total = db.query(Processo).filter(Processo.oculto == False).count()
    urgentes = db.query(Processo).filter(Processo.prioridade == "urgente", Processo.oculto == False).count()
    alta = db.query(Processo).filter(Processo.prioridade == "alta", Processo.oculto == False).count()
    com_tarefa = db.query(Processo).filter(
        Processo.o_que_fazer != None,
        Processo.o_que_fazer != "",
        Processo.oculto == False
    ).count()
    return {
        "total_processos": total,
        "urgentes": urgentes,
        "alta_prioridade": alta,
        "com_tarefa_pendente": com_tarefa,
        "sincronizacao": sincronizacao_status,
    }


@app.post("/api/sincronizar")
async def sincronizar(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    if sincronizacao_status["em_andamento"]:
        raise HTTPException(status_code=409, detail="Sincronização já em andamento.")
    background_tasks.add_task(_executar_sincronizacao, db)
    return {"mensagem": "Sincronização iniciada em segundo plano."}


@app.get("/api/processos", response_model=List[RespostaProcesso])
def listar_processos(
    busca: Optional[str] = Query(None),
    tribunal: Optional[str] = Query(None),
    prioridade: Optional[str] = Query(None),
    situacao: Optional[str] = Query(None),
    ocultos: bool = Query(False),
    ordenar_por: str = Query("atualizado_em"),
    ordem: str = Query("desc"),
    db: Session = Depends(get_db),
):
    q = db.query(Processo)

    if not ocultos:
        q = q.filter(Processo.oculto == False)

    if busca:
        termo = f"%{busca}%"
        q = q.filter(or_(
            Processo.numero.ilike(termo),
            Processo.polo_ativo.ilike(termo),
            Processo.polo_passivo.ilike(termo),
            Processo.vara.ilike(termo),
            Processo.assunto.ilike(termo),
            Processo.classe.ilike(termo),
            Processo.o_que_fazer.ilike(termo),
        ))

    if tribunal:
        q = q.filter(Processo.tribunal.ilike(f"%{tribunal}%"))

    if prioridade:
        q = q.filter(Processo.prioridade == prioridade)

    if situacao:
        q = q.filter(Processo.situacao.ilike(f"%{situacao}%"))

    campo_ordem = getattr(Processo, ordenar_por, Processo.atualizado_em)
    if ordem == "desc":
        q = q.order_by(campo_ordem.desc())
    else:
        q = q.order_by(campo_ordem.asc())

    return q.all()


@app.get("/api/processos/{numero}", response_model=RespostaProcesso)
def detalhe_processo(numero: str, db: Session = Depends(get_db)):
    processo = db.query(Processo).filter(Processo.numero == numero).first()
    if not processo:
        raise HTTPException(status_code=404, detail="Processo não encontrado.")
    return processo


@app.patch("/api/processos/{numero}", response_model=RespostaProcesso)
def atualizar_processo(numero: str, dados: AtualizacaoProcesso, db: Session = Depends(get_db)):
    processo = db.query(Processo).filter(Processo.numero == numero).first()
    if not processo:
        raise HTTPException(status_code=404, detail="Processo não encontrado.")

    if dados.o_que_fazer is not None:
        processo.o_que_fazer = dados.o_que_fazer
    if dados.prioridade is not None:
        processo.prioridade = dados.prioridade
    if dados.oculto is not None:
        processo.oculto = dados.oculto

    processo.atualizado_em = datetime.utcnow()
    db.commit()
    db.refresh(processo)
    return processo


@app.get("/api/tribunais")
def listar_tribunais(db: Session = Depends(get_db)):
    tribunais = db.query(Processo.tribunal).distinct().filter(Processo.tribunal != None).all()
    return sorted([t[0] for t in tribunais if t[0]])


@app.get("/api/exportar/excel")
def exportar_excel(db: Session = Depends(get_db)):
    processos = db.query(Processo).filter(Processo.oculto == False).order_by(
        Processo.prioridade.asc(), Processo.atualizado_em.desc()
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"

    CORES_PRIORIDADE = {
        "urgente": "FF4444",
        "alta": "FF8C00",
        "normal": "2196F3",
        "baixa": "9E9E9E",
    }

    # Cabeçalho
    cabecalhos = [
        "Nº do Processo", "Tribunal", "Vara/Órgão", "Classe", "Assunto",
        "Data Distribuição", "Polo Ativo", "Polo Passivo",
        "Valor da Causa (R$)", "Situação Atual",
        "Último Movimento", "Data Último Mov.", "Prioridade",
        "O que fazer / Anotações", "OAB de Origem", "Atualizado em"
    ]

    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, titulo in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = borda

    ws.row_dimensions[1].height = 40

    # Dados
    for row_idx, p in enumerate(processos, 2):
        valores = [
            p.numero,
            p.tribunal or "",
            p.vara or "",
            p.classe or "",
            p.assunto or "",
            p.data_distribuicao or "",
            p.polo_ativo or "",
            p.polo_passivo or "",
            p.valor_causa,
            p.situacao or "",
            p.ultimo_movimento or "",
            p.data_ultimo_movimento or "",
            (p.prioridade or "normal").upper(),
            p.o_que_fazer or "",
            p.fonte_oab or "",
            p.atualizado_em.strftime("%d/%m/%Y %H:%M") if p.atualizado_em else "",
        ]

        cor_pri = CORES_PRIORIDADE.get(p.prioridade or "normal", "FFFFFF")
        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")

        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = borda
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col_idx == 13:  # Prioridade
                cell.fill = PatternFill(start_color=cor_pri, end_color=cor_pri, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 14:  # O que fazer
                cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                cell.font = Font(italic=True)
            else:
                cell.fill = row_fill

    # Larguras das colunas
    larguras = [25, 35, 30, 25, 30, 18, 35, 35, 18, 25, 40, 18, 14, 45, 18, 18]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"

    # Rodapé
    linha_rodape = len(processos) + 3
    ws.cell(row=linha_rodape, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
    ws.cell(row=linha_rodape + 1, column=1, value="Dra. Daniele de Assis Santiago Cabral — OAB/RR 617 | OAB/AM 1.404-A")
    ws.cell(row=linha_rodape + 1, column=1).font = Font(bold=True, italic=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"processos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
    )


# Montar frontend estático (deve ser o último)
montar_frontend()
